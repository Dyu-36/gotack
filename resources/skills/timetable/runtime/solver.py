#!/usr/bin/env python3
"""Declarative CP-SAT timetable runner used by the bundled timetable skill.

Input is a JSON problem. The agent translates confirmed business requirements
into this schema; it must not generate a search algorithm. The runner owns
model construction, assumptions/unsat-core reporting, progress, validation,
and optional XLSX export.
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import threading
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

from ortools.sat.python import cp_model

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

LIMIT_SECONDS = 90.0
HEARTBEAT_SECONDS = 5.0


def emit(payload: dict[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), flush=True)


def load_problem(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        value = json.load(handle)
    if not isinstance(value, dict):
        raise ValueError("problem JSON must be an object")
    return value


def description(item: dict[str, Any], fallback: str) -> str:
    return str(item.get("description") or item.get("id") or fallback)


def slot_key(slot: dict[str, Any]) -> tuple[str, str, int]:
    return (str(slot.get("day", "")), str(slot.get("session", "")), int(slot.get("period", 0)))


def slot_matches(slot: dict[str, Any], selector: dict[str, Any]) -> bool:
    for key in ("day", "session", "period"):
        if key in selector and selector[key] is not None and str(slot.get(key)) != str(selector[key]):
            return False
    return True


def req_matches(req: dict[str, Any], selector: dict[str, Any]) -> bool:
    aliases = {"class": "class", "subject": "subject", "teacher": "teacher", "requirement_id": "id"}
    for source, target in aliases.items():
        if source in selector and selector[source] not in (None, "") and str(req.get(target, "")) != str(selector[source]):
            return False
    return True


def normalize(problem: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    slots = list(problem.get("slots") or [])
    reqs = list(problem.get("requirements") or [])
    hard = list(problem.get("hard_constraints") or [])
    soft = list(problem.get("soft_constraints") or [])
    for idx, slot in enumerate(slots):
        slot.setdefault("id", f"slot:{idx}")
        slot["period"] = int(slot.get("period", 0))
    for idx, req in enumerate(reqs):
        req.setdefault("id", f"requirement:{idx}")
        req["periods"] = int(req.get("periods", 0))
    return slots, reqs, hard, soft


def validate_input(slots: list[dict[str, Any]], reqs: list[dict[str, Any]], hard: list[dict[str, Any]], soft: list[dict[str, Any]]) -> list[str]:
    errors: list[str] = []
    if not slots:
        errors.append("Dữ liệu không có ô thời khóa biểu khả dụng.")
    if not reqs:
        errors.append("Dữ liệu không có phân công lớp/môn/giáo viên cần xếp.")
    seen_slots: set[tuple[str, str, int]] = set()
    for slot in slots:
        key = slot_key(slot)
        if not key[0] or key[2] <= 0:
            errors.append(f"Ô lịch {slot.get('id')} thiếu thứ hoặc số tiết hợp lệ.")
        if key in seen_slots:
            errors.append(f"Ô lịch bị trùng: {key[0]} {key[1]} tiết {key[2]}.")
        seen_slots.add(key)
    for req in reqs:
        missing = [field for field in ("class", "subject", "teacher") if not str(req.get(field, "")).strip()]
        if missing:
            errors.append(f"Phân công {req.get('id')} thiếu {', '.join(missing)}.")
        if req.get("periods", 0) <= 0:
            errors.append(f"Phân công {req.get('id')} phải có số tiết lớn hơn 0.")
    supported_hard = {
        "teacher_allowed_days", "teacher_forbidden_days", "class_allowed_days", "class_forbidden_days",
        "teacher_workday_count", "teacher_max_periods_per_day", "fixed_slot", "forbidden_slot",
    }
    supported_soft = {
        "teacher_preferred_days", "teacher_avoid_days", "class_preferred_days", "class_avoid_days",
        "preferred_slot", "avoid_slot", "teacher_workday_count", "teacher_max_periods_per_day",
    }
    for item in hard:
        if item.get("type") not in supported_hard:
            errors.append(f"Hard constraint {item.get('id', '<không id>')} có type không hỗ trợ: {item.get('type')}.")
    for item in soft:
        if item.get("type") not in supported_soft:
            errors.append(f"Soft constraint {item.get('id', '<không id>')} có type không hỗ trợ: {item.get('type')}.")
        try:
            if int(item.get("weight", 1)) <= 0:
                errors.append(f"Soft constraint {item.get('id', '<không id>')} phải có weight > 0.")
        except (TypeError, ValueError):
            errors.append(f"Soft constraint {item.get('id', '<không id>')} có weight không hợp lệ.")
    return errors


def fast_conflicts(slots: list[dict[str, Any]], reqs: list[dict[str, Any]], hard: list[dict[str, Any]]) -> list[dict[str, str]]:
    conflicts: list[dict[str, str]] = []
    class_periods: dict[str, int] = defaultdict(int)
    teacher_periods: dict[str, int] = defaultdict(int)
    for req in reqs:
        class_periods[str(req["class"])] += int(req["periods"])
        teacher_periods[str(req["teacher"])] += int(req["periods"])
    for klass, count in class_periods.items():
        if count > len(slots):
            conflicts.append({"id": f"class:{klass}:capacity", "description": f"Lớp {klass} cần {count} tiết nhưng chỉ có {len(slots)} ô lịch khả dụng."})

    allowed_days: dict[str, set[str]] = {}
    forbidden_days: dict[str, set[str]] = defaultdict(set)
    workdays: dict[str, int] = {}
    max_per_day: dict[str, int] = {}
    for item in hard:
        kind = item.get("type")
        teacher = str(item.get("teacher", ""))
        if kind == "teacher_allowed_days" and teacher:
            allowed_days[teacher] = {str(v) for v in item.get("days", [])}
        elif kind == "teacher_forbidden_days" and teacher:
            forbidden_days[teacher].update(str(v) for v in item.get("days", []))
        elif kind == "teacher_workday_count" and teacher:
            workdays[teacher] = int(item.get("count", 0))
        elif kind == "teacher_max_periods_per_day" and teacher:
            max_per_day[teacher] = int(item.get("count", 0))

    slots_by_day: dict[str, int] = defaultdict(int)
    for slot in slots:
        slots_by_day[str(slot["day"])] += 1
    all_days = set(slots_by_day)
    for teacher, count in teacher_periods.items():
        days = allowed_days.get(teacher, set(all_days)) - forbidden_days.get(teacher, set())
        capacity = sum(slots_by_day[d] for d in days)
        if teacher in max_per_day:
            capacity = min(capacity, len(days) * max_per_day[teacher])
        if count > capacity:
            conflicts.append({"id": f"teacher:{teacher}:capacity", "description": f"Giáo viên {teacher} cần {count} tiết nhưng các ngày/giới hạn hiện tại chỉ cho tối đa {capacity} tiết."})
        if teacher in workdays:
            required_days = workdays[teacher]
            if required_days > len(days) or required_days > count or required_days < 0:
                conflicts.append({"id": f"teacher:{teacher}:workday_count", "description": f"Giáo viên {teacher} phải dạy đúng {required_days} ngày nhưng số tiết/ngày khả dụng không tương thích."})

    fixed: list[tuple[dict[str, Any], dict[str, Any], dict[str, Any]]] = []
    for item in hard:
        if item.get("type") != "fixed_slot":
            continue
        candidates = [r for r in reqs if req_matches(r, item)]
        matching_slots = [s for s in slots if slot_matches(s, item)]
        if len(candidates) == 1 and len(matching_slots) == 1:
            fixed.append((item, candidates[0], matching_slots[0]))
    for i, (item_a, req_a, slot_a) in enumerate(fixed):
        for item_b, req_b, slot_b in fixed[i + 1:]:
            if slot_key(slot_a) != slot_key(slot_b):
                continue
            same_class = req_a["class"] == req_b["class"]
            same_teacher = req_a["teacher"] == req_b["teacher"]
            if same_class or same_teacher:
                who = f"lớp {req_a['class']}" if same_class else f"giáo viên {req_a['teacher']}"
                conflicts.append({"id": f"{item_a.get('id')}+{item_b.get('id')}", "description": f"Hai tiết cố định trùng {who} tại {slot_a['day']} {slot_a.get('session', '')} tiết {slot_a['period']}."})
    return conflicts


class Heartbeat:
    def __init__(self, limit_seconds: float) -> None:
        self.start = time.monotonic()
        self.limit_seconds = limit_seconds
        self.solutions = 0
        self.penalty: int | None = None
        self.done = threading.Event()
        self.lock = threading.Lock()

    def update_solution(self, penalty: int | None) -> None:
        with self.lock:
            self.solutions += 1
            self.penalty = penalty
        emit({"type": "progress", "elapsed_seconds": round(time.monotonic() - self.start), "solutions": self.solutions, "penalty": penalty})

    def run(self) -> None:
        while not self.done.wait(HEARTBEAT_SECONDS):
            with self.lock:
                solutions = self.solutions
            emit({"type": "heartbeat", "elapsed_seconds": round(time.monotonic() - self.start), "limit_seconds": int(self.limit_seconds), "solutions": solutions, "state": "SEARCHING" if solutions == 0 else "OPTIMIZING"})


class SolutionProgress(cp_model.CpSolverSolutionCallback):
    def __init__(self, heartbeat: Heartbeat, has_objective: bool, stop_after_first_solution: bool = False) -> None:
        super().__init__()
        self.heartbeat = heartbeat
        self.has_objective = has_objective
        self.stop_after_first_solution = stop_after_first_solution

    def OnSolutionCallback(self) -> None:
        penalty = int(round(self.ObjectiveValue())) if self.has_objective else 0
        self.heartbeat.update_solution(penalty)


def solve(problem: dict[str, Any], problem_path: Path, *, stop_after_first_solution: bool = False) -> dict[str, Any]:
    slots, reqs, hard, soft = normalize(problem)
    errors = validate_input(slots, reqs, hard, soft)
    if errors:
        return {"type": "result", "status": "INPUT_INVALID", "business_result": False, "errors": errors}
    prechecks = fast_conflicts(slots, reqs, hard)
    if prechecks:
        return {"type": "result", "status": "INFEASIBLE", "business_result": True, "hard_constraints_satisfied": False, "conflict_source": "precheck", "hard_conflicts": prechecks}

    model = cp_model.CpModel()
    x: dict[tuple[int, int], cp_model.IntVar] = {}
    for ri, req in enumerate(reqs):
        for si, slot in enumerate(slots):
            x[ri, si] = model.NewBoolVar(f"x_{ri}_{si}")

    assumption_map: dict[int, dict[str, str]] = {}

    def assumption(item_id: str, desc: str) -> cp_model.IntVar:
        lit = model.NewBoolVar(f"assume_{len(assumption_map)}")
        model.AddAssumption(lit)
        assumption_map[lit.Index()] = {"id": item_id, "description": desc}
        return lit

    for ri, req in enumerate(reqs):
        item_id = f"requirement:{req['id']}:required_periods"
        lit = assumption(item_id, f"{req['class']} - {req['subject']} ({req['teacher']}) phải đủ {req['periods']} tiết.")
        model.Add(sum(x[ri, si] for si in range(len(slots))) == int(req["periods"])).OnlyEnforceIf(lit)

    by_class: dict[str, list[int]] = defaultdict(list)
    by_teacher: dict[str, list[int]] = defaultdict(list)
    for ri, req in enumerate(reqs):
        by_class[str(req["class"])].append(ri)
        by_teacher[str(req["teacher"])].append(ri)
    for klass, ris in by_class.items():
        lit = assumption(f"class:{klass}:no_overlap", f"Lớp {klass} không được học hai môn cùng một ô lịch.")
        for si in range(len(slots)):
            model.Add(sum(x[ri, si] for ri in ris) <= 1).OnlyEnforceIf(lit)
    for teacher, ris in by_teacher.items():
        lit = assumption(f"teacher:{teacher}:no_overlap", f"Giáo viên {teacher} không được dạy hai lớp cùng một ô lịch.")
        for si in range(len(slots)):
            model.Add(sum(x[ri, si] for ri in ris) <= 1).OnlyEnforceIf(lit)

    days = list(dict.fromkeys(str(slot["day"]) for slot in slots))
    slots_for_day = {day: [si for si, slot in enumerate(slots) if str(slot["day"]) == day] for day in days}
    teacher_day_used: dict[tuple[str, str], cp_model.IntVar] = {}

    def day_used(teacher: str, day: str) -> cp_model.IntVar:
        key = (teacher, day)
        if key in teacher_day_used:
            return teacher_day_used[key]
        var = model.NewBoolVar(f"day_{len(teacher_day_used)}")
        relevant = [x[ri, si] for ri in by_teacher.get(teacher, []) for si in slots_for_day.get(day, [])]
        if relevant:
            model.Add(sum(relevant) >= var)
            model.Add(sum(relevant) <= len(relevant) * var)
        else:
            model.Add(var == 0)
        teacher_day_used[key] = var
        return var

    for idx, item in enumerate(hard):
        kind = str(item.get("type"))
        item_id = str(item.get("id") or f"hard:{idx}:{kind}")
        desc = description(item, item_id)
        lit = assumption(item_id, desc)
        matching = [ri for ri, req in enumerate(reqs) if req_matches(req, item)]
        if kind in {"teacher_allowed_days", "teacher_forbidden_days", "class_allowed_days", "class_forbidden_days"}:
            allowed = {str(v) for v in item.get("days", [])}
            is_allowed = kind.endswith("allowed_days")
            for ri in matching:
                for si, slot in enumerate(slots):
                    day_ok = str(slot["day"]) in allowed
                    if (is_allowed and not day_ok) or (not is_allowed and day_ok):
                        model.Add(x[ri, si] == 0).OnlyEnforceIf(lit)
        elif kind == "teacher_workday_count":
            teacher = str(item.get("teacher", ""))
            count = int(item.get("count", 0))
            model.Add(sum(day_used(teacher, day) for day in days) == count).OnlyEnforceIf(lit)
        elif kind == "teacher_max_periods_per_day":
            teacher = str(item.get("teacher", ""))
            count = int(item.get("count", 0))
            for day in days:
                vars_for_day = [x[ri, si] for ri in by_teacher.get(teacher, []) for si in slots_for_day[day]]
                if vars_for_day:
                    model.Add(sum(vars_for_day) <= count).OnlyEnforceIf(lit)
        elif kind in {"fixed_slot", "forbidden_slot"}:
            matching_slots = [si for si, slot in enumerate(slots) if slot_matches(slot, item)]
            vars_selected = [x[ri, si] for ri in matching for si in matching_slots]
            if kind == "fixed_slot":
                model.Add(sum(vars_selected) == 1).OnlyEnforceIf(lit)
            else:
                model.Add(sum(vars_selected) == 0).OnlyEnforceIf(lit)

    penalty_terms: list[cp_model.LinearExpr] = []
    soft_metrics: list[dict[str, Any]] = []
    for idx, item in enumerate(soft):
        kind = str(item.get("type"))
        item_id = str(item.get("id") or f"soft:{idx}:{kind}")
        desc = description(item, item_id)
        weight = int(item.get("weight", 1))
        matching = [ri for ri, req in enumerate(reqs) if req_matches(req, item)]
        metric_vars: list[cp_model.IntVar] = []
        if kind in {"teacher_preferred_days", "class_preferred_days", "teacher_avoid_days", "class_avoid_days"}:
            target_days = {str(v) for v in item.get("days", [])}
            prefer = kind.endswith("preferred_days")
            for ri in matching:
                for si, slot in enumerate(slots):
                    in_target = str(slot["day"]) in target_days
                    if (prefer and not in_target) or (not prefer and in_target):
                        metric_vars.append(x[ri, si])
        elif kind in {"preferred_slot", "avoid_slot"}:
            prefer = kind == "preferred_slot"
            for ri in matching:
                for si, slot in enumerate(slots):
                    selected = slot_matches(slot, item)
                    if (prefer and not selected) or (not prefer and selected):
                        metric_vars.append(x[ri, si])
        elif kind == "teacher_workday_count":
            teacher = str(item.get("teacher", ""))
            target = int(item.get("count", 0))
            actual = sum(day_used(teacher, day) for day in days)
            diff = model.NewIntVar(0, len(days), f"soft_workdays_{idx}")
            model.AddAbsEquality(diff, actual - target)
            metric_vars.append(diff)
        elif kind == "teacher_max_periods_per_day":
            teacher = str(item.get("teacher", ""))
            target = int(item.get("count", 0))
            for day in days:
                vars_for_day = [x[ri, si] for ri in by_teacher.get(teacher, []) for si in slots_for_day[day]]
                if not vars_for_day:
                    continue
                excess = model.NewIntVar(0, len(vars_for_day), f"soft_excess_{idx}_{len(metric_vars)}")
                model.Add(excess >= sum(vars_for_day) - target)
                metric_vars.append(excess)
        if metric_vars:
            penalty_terms.extend(var * weight for var in metric_vars)
        soft_metrics.append({"id": item_id, "description": desc, "weight": weight, "vars": metric_vars})

    if penalty_terms:
        model.Minimize(sum(penalty_terms))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = LIMIT_SECONDS
    solver.parameters.num_search_workers = min(4, os.cpu_count() or 1)
    solver.parameters.random_seed = 42
    solver.parameters.stop_after_first_solution = stop_after_first_solution
    if stop_after_first_solution:
        # Test hook for the "budget ended after an incumbent" path: avoid
        # presolve proving optimality before the first-solution stop fires.
        solver.parameters.cp_model_presolve = False

    heartbeat = Heartbeat(LIMIT_SECONDS)
    thread = threading.Thread(target=heartbeat.run, name="timetable-heartbeat", daemon=True)
    thread.start()
    try:
        status = solver.Solve(model, SolutionProgress(heartbeat, bool(penalty_terms), stop_after_first_solution))
    finally:
        heartbeat.done.set()
        thread.join(timeout=1.0)

    status_name = solver.StatusName(status)
    elapsed = round(time.monotonic() - heartbeat.start)
    if status == cp_model.INFEASIBLE:
        core = solver.SufficientAssumptionsForInfeasibility()
        conflicts: list[dict[str, str]] = []
        seen: set[str] = set()
        for literal in core:
            index = int(literal)
            if index < 0:
                index = -index - 1
            item = assumption_map.get(index)
            if item and item["id"] not in seen:
                conflicts.append(item)
                seen.add(item["id"])
        return {"type": "result", "status": "INFEASIBLE", "business_result": True, "hard_constraints_satisfied": False, "elapsed_seconds": elapsed, "conflict_source": "unsat_core", "hard_conflicts": conflicts}
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return {"type": "result", "status": status_name or "UNKNOWN", "business_result": False, "hard_constraints_satisfied": False, "elapsed_seconds": elapsed}

    schedule: list[dict[str, Any]] = []
    for ri, req in enumerate(reqs):
        for si, slot in enumerate(slots):
            if solver.Value(x[ri, si]):
                schedule.append({"day": slot["day"], "session": slot.get("session", ""), "period": slot["period"], "class": req["class"], "subject": req["subject"], "teacher": req["teacher"], "requirement_id": req["id"]})
    order = {slot_key(slot): i for i, slot in enumerate(slots)}
    schedule.sort(key=lambda row: (order.get((str(row["day"]), str(row.get("session", "")), int(row["period"])), 10**9), str(row["class"]), str(row["subject"])))

    validation_errors = validate_schedule(schedule, slots, reqs, hard)
    if validation_errors:
        return {"type": "result", "status": "RUNTIME_ERROR", "business_result": False, "hard_constraints_satisfied": False, "elapsed_seconds": elapsed, "errors": validation_errors}

    violations: list[dict[str, Any]] = []
    total_penalty = 0
    for metric in soft_metrics:
        amount = sum(int(solver.Value(var)) for var in metric["vars"])
        if amount > 0:
            penalty = amount * int(metric["weight"])
            total_penalty += penalty
            violations.append({"id": metric["id"], "description": metric["description"], "violation": amount, "weight": metric["weight"], "penalty": penalty})

    output_xlsx = problem.get("output_xlsx")
    if output_xlsx:
        write_xlsx(schedule, problem, problem_path)

    return {"type": "result", "status": "OPTIMAL" if status == cp_model.OPTIMAL else "FEASIBLE", "business_result": True, "hard_constraints_satisfied": True, "elapsed_seconds": elapsed, "total_penalty": total_penalty, "soft_violations": violations, "schedule": schedule, "output_xlsx": str(resolve_path(problem_path, output_xlsx)) if output_xlsx else None}


def validate_schedule(schedule: list[dict[str, Any]], slots: list[dict[str, Any]], reqs: list[dict[str, Any]], hard: list[dict[str, Any]]) -> list[str]:
    errors: list[str] = []
    valid_slots = {slot_key(slot) for slot in slots}
    by_req: dict[str, int] = defaultdict(int)
    class_slot: set[tuple[str, tuple[str, str, int]]] = set()
    teacher_slot: set[tuple[str, tuple[str, str, int]]] = set()
    for row in schedule:
        key = (str(row["day"]), str(row.get("session", "")), int(row["period"]))
        if key not in valid_slots:
            errors.append(f"Lịch chứa ô không tồn tại: {key}.")
        by_req[str(row["requirement_id"])] += 1
        ck = (str(row["class"]), key)
        tk = (str(row["teacher"]), key)
        if ck in class_slot:
            errors.append(f"Lớp {row['class']} bị trùng tại {key}.")
        if tk in teacher_slot:
            errors.append(f"Giáo viên {row['teacher']} bị trùng tại {key}.")
        class_slot.add(ck)
        teacher_slot.add(tk)
    for req in reqs:
        if by_req[str(req["id"])] != int(req["periods"]):
            errors.append(f"{req['class']} - {req['subject']} ({req['teacher']}) có {by_req[str(req['id'])]} tiết, cần {req['periods']}.")

    rows_by_teacher: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in schedule:
        rows_by_teacher[str(row["teacher"])].append(row)
    for item in hard:
        kind = str(item.get("type"))
        matching = [row for row in schedule if req_matches(row | {"id": row.get("requirement_id", "")}, item)]
        desc = description(item, str(item.get("id", kind)))
        days = {str(v) for v in item.get("days", [])}
        if kind.endswith("allowed_days") and any(str(row["day"]) not in days for row in matching):
            errors.append(desc)
        elif kind.endswith("forbidden_days") and any(str(row["day"]) in days for row in matching):
            errors.append(desc)
        elif kind == "fixed_slot":
            count = sum(1 for row in matching if slot_matches(row, item))
            if count != 1:
                errors.append(desc)
        elif kind == "forbidden_slot" and any(slot_matches(row, item) for row in matching):
            errors.append(desc)
        elif kind == "teacher_workday_count":
            teacher = str(item.get("teacher", ""))
            actual = len({str(row["day"]) for row in rows_by_teacher.get(teacher, [])})
            if actual != int(item.get("count", 0)):
                errors.append(desc)
        elif kind == "teacher_max_periods_per_day":
            teacher = str(item.get("teacher", ""))
            counts: dict[str, int] = defaultdict(int)
            for row in rows_by_teacher.get(teacher, []):
                counts[str(row["day"])] += 1
            if any(value > int(item.get("count", 0)) for value in counts.values()):
                errors.append(desc)
    return errors


def resolve_path(problem_path: Path, value: str | os.PathLike[str]) -> Path:
    path = Path(value)
    return path if path.is_absolute() else (problem_path.parent / path).resolve()


def write_xlsx(schedule: list[dict[str, Any]], problem: dict[str, Any], problem_path: Path) -> None:
    from openpyxl import load_workbook

    template_value = problem.get("template_xlsx")
    template = resolve_path(problem_path, template_value) if template_value else Path(__file__).resolve().parent.parent / "assets" / "mau-thoi-khoa-bieu.xlsx"
    output = resolve_path(problem_path, str(problem["output_xlsx"]))
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)
    workbook = load_workbook(output)
    if "Dữ liệu" not in workbook.sheetnames:
        raise ValueError("template is missing sheet 'Dữ liệu'")
    sheet = workbook["Dữ liệu"]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for row in schedule:
        sheet.append([row["day"], row.get("session", ""), row["period"], row["class"], row["subject"], row["teacher"]])
    workbook.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Solve a declarative school timetable with OR-Tools CP-SAT")
    parser.add_argument("problem", help="Path to problem.json")
    args = parser.parse_args()
    problem_path = Path(args.problem).resolve()
    try:
        result = solve(load_problem(problem_path), problem_path)
    except Exception as exc:  # operational failure, never business INFEASIBLE
        emit({"type": "result", "status": "RUNTIME_ERROR", "business_result": False, "hard_constraints_satisfied": False, "error": f"{type(exc).__name__}: {exc}"})
        return 2
    emit(result)
    return 0 if result.get("business_result") else 2


if __name__ == "__main__":
    raise SystemExit(main())
