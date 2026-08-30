#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script xuất bảng thời khóa biểu ra tệp Excel chuẩn từ schedule.json.
Sử dụng thư viện openpyxl để tạo định dạng A4 Landscape chuẩn trường học.
"""

import json 
import sys 
from collections import Counter ,defaultdict 
from pathlib import Path 

if hasattr (sys .stdout ,"reconfigure"):
    sys .stdout .reconfigure (encoding ="utf-8")
if hasattr (sys .stderr ,"reconfigure"):
    sys .stderr .reconfigure (encoding ="utf-8")

try :
    from openpyxl import Workbook 
    from openpyxl .styles import Alignment ,Border ,Font ,PatternFill ,Side 
    from openpyxl .utils import get_column_letter 
    from openpyxl .worksheet .properties import PageSetupProperties 
except ImportError :
    sys .stderr .write ("Lỗi: Thư viện 'openpyxl' chưa được cài đặt. Hãy chạy: pip install openpyxl\n")
    sys .exit (1 )

FONT_NAME ="Arial"
FONT_SIZE =10 
SHEET_TITLE ="Thời khóa biểu"

THIN_BORDER =Side (border_style ="thin",color ="000000")
MEDIUM_BORDER =Side (border_style ="medium",color ="000000")
SESSION_ORDER ={"Sáng":1 ,"Chiều":2 ,"Tối":3 }


def _is_text (val ):
    return isinstance (val ,str )and bool (val .strip ())


def _is_int (val ):
    return isinstance (val ,int )and not isinstance (val ,bool )


def _get_day_slots (days ):
    slots =[]
    errors =[]
    if not isinstance (days ,list )or not days :
        errors .append ("Danh sách 'days' không được để trống")
        return slots ,errors 

    for idx ,item in enumerate (days ):
        if not isinstance (item ,dict ):
            errors .append (f"days[{idx }] phải là đối tượng")
            continue 
        day =item .get ("day")
        if not _is_int (day )or day <2 or day >8 :
            errors .append (f"days[{idx }].day phải là số từ 2 đến 8")
            continue 

        sessions =item .get ("sessions")
        if isinstance (sessions ,list )and sessions :
            for s_idx ,sess in enumerate (sessions ):
                name =sess .get ("name","").strip ()if isinstance (sess ,dict )else ""
                periods =sess .get ("periods")if isinstance (sess ,dict )else None 
                if not name or not _is_int (periods )or periods <=0 :
                    errors .append (f"days[{idx }].sessions[{s_idx }] không hợp lệ")
                    continue 
                for p in range (1 ,periods +1 ):
                    slots .append ((day ,name ,p ))
        else :
            periods =item .get ("periods")
            if not _is_int (periods )or periods <=0 :
                errors .append (f"days[{idx }] thiếu cấu hình số tiết hợp lệ")
                continue 
            for p in range (1 ,periods +1 ):
                slots .append ((day ,"",p ))
    return slots ,errors 


def validate (data ):
    errors =[]
    if not isinstance (data ,dict ):
        return ["Dữ liệu phải là JSON object"]

    classes =data .get ("classes")
    if not isinstance (classes ,list )or not classes or not all (_is_text (c )for c in classes ):
        errors .append ("'classes' phải là danh sách tên lớp không rỗng")

    days =data .get ("days")
    day_slots ,day_errors =_get_day_slots (days )
    errors .extend (day_errors )

    assignments =data .get ("assignments")
    if not isinstance (assignments ,list ):
        errors .append ("'assignments' phải là danh sách các tiết học")
        return errors 

    class_set =set (classes )if isinstance (classes ,list )else set ()
    slot_set =set (day_slots )

    by_class =defaultdict (list )
    by_teacher =defaultdict (list )

    for idx ,item in enumerate (assignments ):
        if not isinstance (item ,dict ):
            errors .append (f"assignments[{idx }] không phải đối tượng")
            continue 
        day =item .get ("day")
        session =str (item .get ("session","")).strip ()
        period =item .get ("period")
        cls_name =item .get ("class")
        subject =item .get ("subject")
        teacher =item .get ("teacher")

        if not (_is_int (day )and _is_int (period )and _is_text (cls_name )and _is_text (subject )and _is_text (teacher )):
            errors .append (f"assignments[{idx }] thiếu thông tin bắt buộc")
            continue 

        if (day ,session ,period )not in slot_set :
            errors .append (f"assignments[{idx }] rơi vào khung giờ không tồn tại (Thứ {day }, {session }, Tiết {period })")
            continue 

        if cls_name not in class_set :
            errors .append (f"assignments[{idx }] chứa lớp '{cls_name }' không có trong danh sách 'classes'")
            continue 

        by_class [(day ,session ,period ,cls_name )].append (subject )
        by_teacher [(day ,session ,period ,teacher )].append (cls_name )


    for (d ,s ,p ,c ),subjs in by_class .items ():
        if len (subjs )>1 :
            errors .append (f"Trùng lịch lớp {c } vào Thứ {d } {s } tiết {p } (các môn: {', '.join (subjs )})")


    for (d ,s ,p ,t ),cls_list in by_teacher .items ():
        if len (cls_list )>1 :
            errors .append (f"Trùng lịch giáo viên {t } vào Thứ {d } {s } tiết {p } (dạy các lớp: {', '.join (cls_list )})")

    return errors 


def build_workbook (data ,out_path ):
    classes =data ["classes"]
    days =data ["days"]
    day_slots ,_ =_get_day_slots (days )

    by_slot ={}
    for item in data ["assignments"]:
        session =str (item .get ("session","")).strip ()
        by_slot [(item ["day"],session ,item ["period"],item ["class"])]=item 

    wb =Workbook ()
    ws =wb .active 
    ws .title =SHEET_TITLE 

    base_font =Font (name =FONT_NAME ,size =FONT_SIZE )
    header_font =Font (name =FONT_NAME ,size =FONT_SIZE ,bold =True )
    center_align =Alignment (horizontal ="center",vertical ="center")
    left_align =Alignment (horizontal ="left",vertical ="center",indent =1 )


    header =["Thứ","Tiết"]
    for cls in classes :
        header .append (cls )
        header .append ("GV Dạy")
    ws .append (header )

    last_col =len (header )
    curr_row =2 
    day_boundaries =[]

    grouped_slots =defaultdict (list )
    for d ,s ,p in day_slots :
        grouped_slots [d ].append ((s ,p ))

    for day_item in days :
        d =day_item ["day"]
        start_r =curr_row 
        for s ,p in grouped_slots [d ]:
            p_label =f"{s } {p }".strip ()
            ws .cell (row =curr_row ,column =2 ,value =p_label )
            for idx ,cls in enumerate (classes ):
                col =3 +idx *2 
                item =by_slot .get ((d ,s ,p ,cls ))
                if item :
                    ws .cell (row =curr_row ,column =col ,value =item .get ("subject",""))
                    ws .cell (row =curr_row ,column =col +1 ,value =item .get ("teacher",""))
            curr_row +=1 
        end_r =curr_row -1 
        if end_r >start_r :
            ws .merge_cells (start_row =start_r ,start_column =1 ,end_row =end_r ,end_column =1 )
        ws .cell (row =start_r ,column =1 ,value =f"Thứ {d }"if d <8 else "Chủ Nhật")
        day_boundaries .append (end_r )

    last_row =curr_row -1 
    boundary_set =set (day_boundaries )


    for r in range (1 ,last_row +1 ):
        for c in range (1 ,last_col +1 ):
            cell =ws .cell (row =r ,column =c )
            cell .font =header_font if r ==1 else base_font 
            cell .alignment =center_align if (r ==1 or c <=2 )else left_align 

            top_border =MEDIUM_BORDER if r ==1 else THIN_BORDER 
            bottom_border =MEDIUM_BORDER if (r ==1 or r in boundary_set or r ==last_row )else THIN_BORDER 
            left_border =MEDIUM_BORDER if (c ==1 or c ==3 )else THIN_BORDER 
            right_border =MEDIUM_BORDER if c ==last_col else THIN_BORDER 

            cell .border =Border (top =top_border ,bottom =bottom_border ,left =left_border ,right =right_border )

            if r ==1 :
                cell .fill =PatternFill (start_color ="F2F2F2",end_color ="F2F2F2",fill_type ="solid")


    ws .column_dimensions ["A"].width =10 
    ws .column_dimensions ["B"].width =12 if any (s for _ ,s ,_ in day_slots )else 8 
    for idx in range (len (classes )):
        ws .column_dimensions [get_column_letter (3 +idx *2 )].width =14 
        ws .column_dimensions [get_column_letter (4 +idx *2 )].width =14 


    ws .freeze_panes ="C2"
    ws .print_title_rows ="1:1"
    ws .page_setup .orientation =ws .ORIENTATION_LANDSCAPE 
    ws .page_setup .paperSize =ws .PAPERSIZE_A4 
    ws .page_setup .fitToWidth =1 
    ws .page_setup .fitToHeight =0 
    ws .sheet_properties .pageSetUpPr =PageSetupProperties (fitToPage =True )

    out_path =Path (out_path )
    out_path .parent .mkdir (parents =True ,exist_ok =True )
    wb .save (out_path )
    return str (out_path )


def main ():
    if len (sys .argv )!=3 :
        sys .stderr .write ("Sử dụng: python exporter.py schedule.json thoi-khoa-bieu.xlsx\n")
        return 2 

    src ,dst =sys .argv [1 ],sys .argv [2 ]
    try :
        with open (src ,"r",encoding ="utf-8")as f :
            data =json .load (f )
    except Exception as e :
        sys .stderr .write (f"Lỗi đọc file {src }: {e }\n")
        return 1 

    errors =validate (data )
    if errors :
        sys .stderr .write ("Dữ liệu thời khóa biểu chưa hợp lệ:\n")
        for err in errors :
            sys .stderr .write (f"  • {err }\n")
        return 1 

    build_workbook (data ,dst )
    print (f"Đã xuất thành công: {dst }")
    return 0 


if __name__ =="__main__":
    sys .exit (main ())
